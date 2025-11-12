"""
Excel document parser using openpyxl.
"""
import hashlib
from pathlib import Path
from typing import Dict, Any, List
import openpyxl
from openpyxl.utils import get_column_letter


class ExcelParser:
    """Parse Excel files and extract text content."""
    
    def __init__(self):
        self.supported_extensions = [".xlsx", ".xls"]
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse an Excel file and extract text content.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Dict containing extracted text and metadata
            
        Raises:
            ValueError: If file is not an Excel file or cannot be parsed
        """
        path = Path(file_path)
        
        if path.suffix.lower() not in self.supported_extensions:
            raise ValueError(f"Unsupported file type: {path.suffix}")
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            # Extract metadata
            metadata = {
                "num_sheets": len(workbook.sheetnames),
                "sheet_names": workbook.sheetnames,
                "file_name": path.name,
                "file_size": path.stat().st_size,
            }
            
            # Extract text from all sheets
            sheets_content = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_text = self._extract_sheet_text(sheet, sheet_name)
                if sheet_text:
                    sheets_content.append(sheet_text)
            
            full_text = "\n\n".join(sheets_content)
            
            # Calculate file hash for deduplication
            file_hash = self._calculate_file_hash(file_path)
            
            return {
                "text": full_text,
                "metadata": metadata,
                "file_hash": file_hash,
                "success": True,
                "error": None
            }
            
        except Exception as e:
            return {
                "text": "",
                "metadata": {},
                "file_hash": "",
                "success": False,
                "error": f"Failed to parse Excel file: {str(e)}"
            }
    
    def _extract_sheet_text(self, sheet, sheet_name: str) -> str:
        """Extract text from a single Excel sheet."""
        lines = [f"=== Sheet: {sheet_name} ==="]
        
        # Get max row and column
        max_row = sheet.max_row
        max_col = sheet.max_column
        
        if max_row == 0 or max_col == 0:
            return ""
        
        # Extract headers (first row)
        headers = []
        for col in range(1, max_col + 1):
            cell_value = sheet.cell(row=1, column=col).value
            headers.append(str(cell_value) if cell_value else "")
        
        lines.append("Headers: " + " | ".join(headers))
        lines.append("")
        
        # Extract data rows
        for row in range(2, min(max_row + 1, 1000)):  # Limit to 1000 rows
            row_data = []
            for col in range(1, max_col + 1):
                cell_value = sheet.cell(row=row, column=col).value
                row_data.append(str(cell_value) if cell_value else "")
            
            # Only add non-empty rows
            if any(val.strip() for val in row_data):
                lines.append(" | ".join(row_data))
        
        return "\n".join(lines)
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of file for deduplication."""
        sha256_hash = hashlib.sha256()
        with open(file_path, "rb") as f:
            for byte_block in iter(lambda: f.read(4096), b""):
                sha256_hash.update(byte_block)
        return sha256_hash.hexdigest()
    
    def validate_file(self, file_path: str, max_size_mb: int = 50) -> Dict[str, Any]:
        """
        Validate Excel file before parsing.
        
        Args:
            file_path: Path to the Excel file
            max_size_mb: Maximum file size in MB
            
        Returns:
            Dict with validation result
        """
        path = Path(file_path)
        
        if not path.exists():
            return {"valid": False, "error": "File not found"}
        
        if path.suffix.lower() not in self.supported_extensions:
            return {"valid": False, "error": f"Invalid file type: {path.suffix}"}
        
        file_size_mb = path.stat().st_size / (1024 * 1024)
        if file_size_mb > max_size_mb:
            return {"valid": False, "error": f"File too large: {file_size_mb:.2f}MB (max: {max_size_mb}MB)"}
        
        return {"valid": True, "error": None}
